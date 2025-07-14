import pandas as pd
import warnings
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Alignment, Protection
from openpyxl.worksheet.worksheet import Worksheet

# template_path = r"C:\Users\hansen\Desktop\Python Scripts\LV_Data_Assorter\Data\VORLAGE_XXX_Budgetbildung auf Grundlage Kobe_HP.xlsx"
# raw_path = r"C:\Users\hansen\Desktop\Python Scripts\LV_Data_Assorter\Data\Kostenermittlung nach LB mit Rezepturen_Manne.xlsx"

warnings.filterwarnings("ignore", message="Workbook contains no default style")

def load_and_clean_data(raw_path: str) -> pd.DataFrame:
    df_raw = pd.read_excel(raw_path, header=10)
    df_raw = df_raw.dropna(axis=1, how='all').dropna(axis=0, how='all').reset_index(drop=True)
    df_raw = df_raw.drop(df_raw.columns[[3,4,5,9,10,11]], axis=1)
    df_raw['Leistungsbereich'] = pd.NA
    df_raw.columns = ['LB','KGR','Bezeichnung','Menge','ME','EP','GB','Leistungsbereich']
    df_raw = df_raw[~(df_raw['LB'].isna() & df_raw['KGR'].isna())].reset_index(drop=True)
    return df_raw

def fill_leistungsbereich_and_kgr(df_raw: pd.DataFrame):
    # Fill Leistungsbereich
    col_source = df_raw.columns[0]
    col_target = df_raw.columns[-1]
    current_value = None
    leistungsbereich_liste = []

    for idx, row in df_raw.iterrows():
        value = row[col_source]
        if pd.notna(value) and not str(value).startswith("0"):
            continue
        if pd.notna(value):
            current_value = value
            bezeichnung = row[df_raw.columns[2]]
            leistungsbereich_liste.append([current_value, bezeichnung])
        df_raw.at[idx, col_target] = current_value

    df_leistungsbereiche = pd.DataFrame(leistungsbereich_liste, columns=['Leistungsbereich', 'Bezeichnung'])

    # Fill KGR
    current_value = None
    for idx, row in df_raw.iterrows():
        value = row[col_source]
        if pd.notna(value) and not str(value).startswith("3"):
            continue
        if pd.notna(value):
            current_value = value
        df_raw.at[idx, df_raw.columns[1]] = current_value

    # Clean before insertion
    df_raw = df_raw[df_raw[col_source].apply(lambda x: pd.isna(x))].reset_index(drop=True)
    df_raw = df_raw[df_raw[df_raw.columns[5]] != 0].reset_index(drop=True)
    df_raw.drop(df_raw.columns[0], axis=1, inplace=True)
    df_raw = df_raw.fillna("").infer_objects(copy=False)

    return df_raw, df_leistungsbereiche

def copy_cell_format(source_cell, target_cell):
    # Copy cell styles: font, fill, border, alignment, protection
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.alignment = source_cell.alignment.copy()
        target_cell.protection = source_cell.protection.copy()

def insert_dataframe_with_formatting(ws: Worksheet, start_row: int, df: pd.DataFrame, format_ws: Worksheet):
    n_rows, n_cols = df.shape

    # Insert empty rows
    ws.insert_rows(start_row, amount=n_rows)

    # Copy formatting row from format_ws row 1 to each new row
    for i in range(n_rows):
        for col_idx in range(1, n_cols + 1):
            source_cell = format_ws.cell(row=1, column=col_idx)
            target_cell = ws.cell(row=start_row + i, column=col_idx)
            copy_cell_format(source_cell, target_cell)

    # Write values
    for i, row in enumerate(df.itertuples(index=False), start=start_row):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

def insert_leistungsbereiche(ws: Worksheet, start_row: int, df_leistungsbereiche: pd.DataFrame):
    # Write leistungsbereiche values without formatting
    for i, row in enumerate(df_leistungsbereiche.itertuples(index=False), start=start_row):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

def insert_into_excel(template_path: str, df_raw: pd.DataFrame, df_leistungsbereiche: pd.DataFrame, sheet_name='Übersicht Budgetaufteilung'):
    wb = load_workbook(template_path)
    sheet = wb[sheet_name]
    format_sheet = wb['Vorlage (DO NOT DELETE)']

    insert_dataframe_with_formatting(sheet, start_row=4, df=df_raw, format_ws=format_sheet)
    insert_leistungsbereiche(format_sheet, start_row=5, df_leistungsbereiche=df_leistungsbereiche)

    wb.save(template_path)
    wb.close()

    wb.save(template_path)
    wb.close()

def run_processing(template_path: str, raw_path: str):
    df_raw = load_and_clean_data(raw_path)
    df_raw, df_leistungsbereiche = fill_leistungsbereich_and_kgr(df_raw)
    insert_into_excel(template_path, df_raw, df_leistungsbereiche)
