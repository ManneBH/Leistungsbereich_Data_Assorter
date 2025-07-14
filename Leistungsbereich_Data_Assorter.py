import streamlit as st
import pandas as pd
import warnings
import io
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle

# Nulstil hele session_state ved første kørsel
if "initialized" not in st.session_state:
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.session_state.initialized = True

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 1000)
pd.set_option('display.max_colwidth', None)
warnings.filterwarnings("ignore", message="Workbook contains no default style")

# Titel og inputfelter (vises kun hvis ikke færdig)
if not st.session_state.get("completed", False):
    st.title("Vorlage-Datei hochladen")
    template_file = st.file_uploader("Upload der VORLAGE-Datei", type=["xlsx"], key="template_file")

    st.title("Quelldatei hochladen")
    raw_file = st.file_uploader("Upload der DATEN-Datei", type=["xlsx"], key="raw_file")

    if template_file and raw_file:
        if st.button("Starte Verarbeitung"):
            st.session_state.run_process = True
    else:
        st.stop()
else:
    st.title("Fertig!")
    st.text("Made by: Manne Bach Hansen")
    if st.button("Neustarten"):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.experimental_rerun()
    st.stop()

# Stop hvis ikke brugeren har trykket knappen
if not st.session_state.get("run_process"):
    st.stop()

# ========== PROCESSEN STARTER HER ==========

try:
    df_template = pd.read_excel(template_file, sheet_name=0)
    df_raw = pd.read_excel(raw_file, header=10)
except Exception as e:
    st.error(f"Fehler beim Einlesen der Dateien: {e}")
    st.stop()

df_raw = df_raw.dropna(axis=1, how='all').dropna(axis=0, how='all').reset_index(drop=True)
df_raw = df_raw.drop(df_raw.columns[[3, 4, 5, 9, 10, 11]], axis=1)
df_raw['Leistungsbereich'] = pd.NA
df_raw.columns = ['LB', 'KGR', 'Bezeichnung', 'Menge', 'ME', 'EP', 'GB', 'Leistungsbereich']
df_raw = df_raw[~(df_raw['LB'].isna() & df_raw['KGR'].isna())].reset_index(drop=True)

# Fyld Leistungsbereich
col_source = df_raw.columns[0]
col_target = df_raw.columns[-1]
current_value = None
leistungsbereich_liste = []

for idx, row in df_raw.iterrows():
    value = row[col_source]
    if pd.notna(value) and not str(value).startswith("0"):
        continue
    if pd.notna(value):
        current_value = value
        bezeichnung = row[df_raw.columns[2]]
        leistungsbereich_liste.append([current_value, bezeichnung])
    df_raw.at[idx, col_target] = current_value

df_leistungsbereiche = pd.DataFrame(leistungsbereich_liste, columns=['Leistungsbereich', 'Bezeichnung'])

# Fyld KGR
current_value = None
for idx, row in df_raw.iterrows():
    value = row[col_source]
    if pd.notna(value) and not str(value).startswith("3"):
        continue
    if pd.notna(value):
        current_value = value
    df_raw.at[idx, df_raw.columns[1]] = current_value

# Rens inden indsættelse
df_raw = df_raw[df_raw[col_source].apply(lambda x: pd.isna(x))].reset_index(drop=True)
df_raw = df_raw[df_raw[df_raw.columns[5]] != 0].reset_index(drop=True)
df_raw.drop(df_raw.columns[0], axis=1, inplace=True)
df_raw = df_raw.fillna("").infer_objects(copy=False)

# === Åbn template med openpyxl ===
template_file.seek(0)  # Reset pointer
wb = load_workbook(filename=template_file)
sheet = wb.worksheets[0]

# Fjern alle sammenflettede celler i arket
for merged_range in list(sheet.merged_cells.ranges):
    sheet.unmerge_cells(str(merged_range))
try:
    format_sheet = wb['Vorlage (DO NOT DELETE)']
except KeyError:
    st.error("Die Vorlage muss ein Blatt namens 'Vorlage (DO NOT DELETE)' enthalten.")
    st.stop()

# Indsæt renset df_raw fra række 5 (Excel er 1-indekseret)
start_row = 5

for r_idx, row in enumerate(dataframe_to_rows(df_raw, index=False, header=False), start=start_row):
    for c_idx, value in enumerate(row, start=1):
        cell = sheet.cell(row=r_idx, column=c_idx)
        cell.value = value

# Indsæt df_leistungsbereiche på "Vorlage (DO NOT DELETE)", fx fra række 6
for r_idx, row in enumerate(dataframe_to_rows(df_leistungsbereiche, index=False, header=False), start=6):
    for c_idx, value in enumerate(row, start=1):
        format_sheet.cell(row=r_idx, column=c_idx, value=value)

# Gem resultatet som ny fil og tillad download
output = io.BytesIO()
wb.save(output)
wb.close()
output.seek(0)

st.success("Dateien verarbeitet!")

st.download_button(
    label="⬇️ Ergebnis-Datei herunterladen",
    data=output,
    file_name="Ergebnis_Budgetdatei.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Marker som færdig
st.session_state.completed = True
st.session_state.run_process = False

# Tilføj knap for at genstarte
if st.button("Neustarten"):
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.experimental_rerun()
